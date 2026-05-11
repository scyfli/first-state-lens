"""Load raw artifacts from disk into PipelineInputs for the orchestrator.

This is the S+4 bridge between the source pullers (which persist into
`etl/raw/`) and the orchestrator's `PipelineInputs` dataclass (which
expects pre-parsed Python objects + GeoDataFrames).

Design:
  - The loader is lenient: missing artifacts surface as None pass-throughs
    or empty lists. The geo-stack-dependent loads (TIGER shapefiles, BG
    population join) live behind importorskip and skip when GDAL is
    absent. This lets local Windows dev exercise the non-geo pipeline
    while CI runs the full pipeline.
  - Manifest construction: when a raw artifact is present, the loader
    synthesizes a `SourceEntry` from on-disk metadata (size + sha256 +
    mtime as last_fetched). This is the offline path; the live --pull
    path passes through FetchResult instances directly.
  - No DSB scraper or grantee data: methodology v0.2.0 carries the DGI
    dollar aggregates inside the dashboard scaffold, but the per-grantee
    detail is gated on the DSB canonical URL (carried open question).
    The loader returns an empty grantee list when the DSB JSON is empty
    or missing; the orchestrator handles that cleanly (zero grantees =
    zero geocoded points = food-resources universe still composes from
    SNAP/farmers when those pullers land).

The loader is intentionally pragmatic: it produces a runnable pipeline
from whatever subset of artifacts is available. The deploy gate
(`require_present=True` in the orchestrator) catches the case where the
pipeline ran but produced incomplete outputs.
"""

from __future__ import annotations

import dataclasses
import datetime
import hashlib
import io
import json
from pathlib import Path
from typing import Any, Optional

import yaml

from etl import __version__ as ETL_VERSION
from etl.lib.manifest import Manifest, SourceEntry
from etl.transforms.geocode import GranteeInput
from etl.transforms.merge_food_resources import FoodResource
from etl.transforms.sb254_effective import BlockGroup, TractInput


# ---------------------------------------------------------------------------
# Result containers
# ---------------------------------------------------------------------------


@dataclasses.dataclass
class LoadedRaw:
    """Everything load_raw_artifacts produced from disk.

    The orchestrator's PipelineInputs is constructed from this in a
    follow-on step that decides parameters + manual-reviews path.
    """

    grantees: list[GranteeInput]
    food_resources_raw: list[FoodResource]
    tracts: list[TractInput]
    state_mfi_median: float
    manifest: Manifest

    # Optional pass-through bytes for non-tabular layers
    lila_geojson: Optional[bytes] = None
    sd2_geojson: Optional[bytes] = None
    dart_routes_geojson: Optional[bytes] = None
    acs_demographics_csv: Optional[bytes] = None
    mhc_csv: Optional[bytes] = None

    # Optional geo-stack GeoDataFrames (None when GDAL absent)
    mmg_counties_gdf: Optional[Any] = None
    target_tracts_gdf: Optional[Any] = None
    weights_bg_gdf: Optional[Any] = None

    # Diagnostics surfaced to the CLI
    geo_stack_available: bool = False
    notes: list[str] = dataclasses.field(default_factory=list)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def load_raw_artifacts(raw_dir: Path, parameters: dict) -> LoadedRaw:
    """Inspect `raw_dir` and build a LoadedRaw from whatever's present."""
    manifest = Manifest(etl_version=ETL_VERSION)
    notes: list[str] = []

    # --- DSB grantees ---
    grantees: list[GranteeInput] = []
    dsb_path = raw_dir / "dsb-grants.json"
    if dsb_path.exists():
        try:
            grantees = _load_dsb_grantees(dsb_path)
            _record_source(manifest, "dsb-grants", dsb_path)
            notes.append(f"dsb-grants: {len(grantees)} grantee(s) loaded")
        except Exception as exc:  # noqa: BLE001
            notes.append(f"dsb-grants: load failed ({exc}); skipping")
    else:
        notes.append("dsb-grants: etl/raw/dsb-grants.json not present; running with zero grantees")

    # --- SNAP retailers (S+5) ---
    food_resources_raw: list[FoodResource] = []
    snap_path = raw_dir / "snap-retailers-de.geojson"
    if snap_path.exists():
        try:
            snap_resources = _load_snap_retailers(snap_path)
            food_resources_raw.extend(snap_resources)
            _record_source(manifest, "snap-retailers", snap_path)
            notes.append(f"snap-retailers: {len(snap_resources)} retailer(s) loaded")
        except Exception as exc:  # noqa: BLE001
            notes.append(f"snap-retailers: load failed ({exc}); skipping")
    else:
        notes.append("snap-retailers: snap-retailers-de.geojson not present; food universe excludes SNAP retailers")

    # --- USDA farmers markets (S+5; scaffold path) ---
    fm_path = raw_dir / "usda-farmers-markets-de.json"
    if fm_path.exists():
        try:
            fm_resources = _load_farmers_markets(fm_path)
            food_resources_raw.extend(fm_resources)
            _record_source(manifest, "usda-farmers-markets", fm_path)
            notes.append(f"usda-farmers-markets: {len(fm_resources)} market(s) loaded")
        except Exception as exc:  # noqa: BLE001
            notes.append(f"usda-farmers-markets: load failed ({exc}); skipping")
    else:
        notes.append("usda-farmers-markets: usda-farmers-markets-de.json not present; food universe excludes farmers markets")

    # --- MMG county data (CSV pass-through path; geo-stack-only for apportion) ---
    mmg_path = raw_dir / "mmg-food-insecurity-counties.csv"
    if mmg_path.exists():
        _record_source(manifest, "mmg-counties", mmg_path)
        notes.append(f"mmg-counties: {mmg_path.stat().st_size} bytes loaded")
    else:
        notes.append("mmg-counties: not present; food-insecurity-tracts.geojson will be empty")

    # --- Pass-through layers (load as bytes when present) ---
    # LILA: pre-clipped geojson takes precedence; if absent, we try to build
    # from the raw xlsx + TIGER geometry below (inside the geo-stack block).
    lila_geojson = _load_passthrough(raw_dir / "usda-lila-de-clipped.geojson", manifest, "usda-lila-clipped", notes)
    sd2_geojson = _load_passthrough(raw_dir / "firstmap-sd2.geojson", manifest, "firstmap-sd2", notes)
    # DART routes: pre-built geojson takes precedence; if absent, build LineStrings
    # from the GTFS zip's shapes.txt (no geopandas required — plain CSV → GeoJSON).
    dart_routes_geojson = _load_passthrough(raw_dir / "dart-routes.geojson", manifest, "dart-routes", notes)
    if dart_routes_geojson is None:
        gtfs_zip = raw_dir / "dart-gtfs.zip"
        if gtfs_zip.exists():
            built = _build_dart_routes_geojson_from_gtfs(gtfs_zip, notes)
            if built is not None:
                dart_routes_geojson = built
                _record_source(manifest, "dart-gtfs-zip", gtfs_zip)
    mhc_csv = _load_passthrough(raw_dir / "mhc-tract-de.csv", manifest, "mhc-tract", notes)

    # --- ACS tract demographics: convert JSON → CSV inline + build per-tract lookup ---
    # Single read; feeds both the CSV pass-through (consumer-facing output)
    # and the demographics_lookup (poverty_rate + mfi joined into TractInput
    # for SB 254 low-income certification).
    acs_tract_path = raw_dir / "acs-tract-de.json"
    acs_demographics_csv: Optional[bytes] = None
    demographics_lookup: dict[str, dict[str, Optional[float]]] = {}
    if acs_tract_path.exists():
        try:
            acs_raw = acs_tract_path.read_bytes()
            acs_demographics_csv = _acs_tract_json_to_csv(acs_raw)
            demographics_lookup = _load_acs_tract_demographics(acs_raw)
            _record_source(manifest, "acs-tract-demographics", acs_tract_path)
            notes.append(
                f"acs-tract-demographics: converted {acs_tract_path.name} → "
                f"{len(acs_demographics_csv)} bytes CSV; "
                f"{len(demographics_lookup)} tract demographics records loaded"
            )
        except Exception as exc:  # noqa: BLE001
            notes.append(f"acs-tract-demographics: convert failed ({exc})")
    else:
        notes.append("acs-tract-demographics: acs-tract-de.json not present; CSV pass-through empty; SB 254 stage will flag tracts as income data missing")

    # --- Geo-stack reads (TIGER + BG join) ---
    tracts: list[TractInput] = []
    target_tracts_gdf = None
    weights_bg_gdf = None
    mmg_counties_gdf = None
    geo_stack = False

    try:
        import geopandas  # noqa: F401  -- the import IS the probe
        geo_stack = True
    except ImportError:
        notes.append("geo-stack: geopandas not installed; skipping shapefile loads + apportionment")

    if geo_stack:
        target_tracts_gdf = _load_tiger_zip(raw_dir / "tiger-tracts-de.zip", manifest, "tiger-tracts", notes)
        weights_bg_gdf = _load_tiger_zip(raw_dir / "tiger-bgs-de.zip", manifest, "tiger-bgs", notes)

        # Build the BG-pop layer (centroid + pop) for sb254 + apportionment weights
        acs_bg_path = raw_dir / "acs-bg-de.json"
        bg_pop_lookup: dict[str, int] = {}
        if acs_bg_path.exists():
            try:
                bg_pop_lookup = _load_acs_bg_pop(acs_bg_path)
                _record_source(manifest, "acs-bg-population", acs_bg_path)
                notes.append(f"acs-bg-population: {len(bg_pop_lookup)} BG records loaded")
            except Exception as exc:  # noqa: BLE001
                notes.append(f"acs-bg-population: load failed ({exc})")
        else:
            notes.append("acs-bg-population: acs-bg-de.json not present; BG-weighted stages will run with zero pop")

        if weights_bg_gdf is not None and bg_pop_lookup:
            weights_bg_gdf = _attach_bg_pop(weights_bg_gdf, bg_pop_lookup)

        # Urban Areas (UAC10) cross-walk — methodology v0.3.0 refinement.
        # Replaces the county-level urbanicity proxy (FIPS 10003 → urban,
        # others → nonurban) with the canonical USDA LILA-aligned
        # per-tract classification. Falls back to the proxy when the
        # UAC zip is absent (graceful degradation).
        urban_areas_gdf = _load_urban_areas_from_national_tiger(
            raw_dir / "tiger-uac-us.zip", manifest, "tiger-uac", notes
        )
        urbanicity_lookup: Optional[dict[str, str]] = None
        if target_tracts_gdf is not None and urban_areas_gdf is not None:
            urbanicity_lookup = _compute_tract_urbanicity_lookup(
                target_tracts_gdf, urban_areas_gdf, notes
            )

        if target_tracts_gdf is not None:
            tracts = _build_tracts_from_gdfs(
                target_tracts_gdf,
                weights_bg_gdf,
                notes,
                demographics_lookup=demographics_lookup,
                urbanicity_lookup=urbanicity_lookup,
            )
            with_demos = sum(1 for t in tracts if t.poverty_rate is not None or t.mfi is not None)
            urban_count = sum(1 for t in tracts if t.urbanicity == "urban")
            notes.append(
                f"tracts: {len(tracts)} tract(s) constructed from TIGER + BG join; "
                f"{with_demos} with ACS demographics joined; "
                f"{urban_count} classified urban"
            )

        # MMG counties GDF (S+5): load national TIGER counties + clip to DE.
        # Joined to mmg-food-insecurity-counties.csv downstream by FIPS.
        # The apportionment stage cleanly skips when mmg_counties_gdf is
        # None OR when mmg-food-insecurity-counties.csv is absent (still
        # gated by the canonical-URL open question).
        mmg_counties_gdf = _load_de_counties_from_national_tiger(
            raw_dir / "tiger-counties-us.zip", manifest, "tiger-counties", notes
        )

        # USDA LILA (S+6 transform): if the pre-clipped geojson wasn't found
        # above AND the raw xlsx + TIGER tract geometry are both available,
        # build the clipped geojson inline. openpyxl is the workbook reader;
        # absent → skip with a diagnostic note.
        lila_xlsx_path = raw_dir / "usda-lila-raw.xlsx"
        if lila_geojson is None and lila_xlsx_path.exists() and target_tracts_gdf is not None:
            built = _build_lila_geojson_from_xlsx(
                lila_xlsx_path, target_tracts_gdf, notes
            )
            if built is not None:
                lila_geojson = built
                _record_source(manifest, "usda-lila-xlsx", lila_xlsx_path)

    state_mfi_median = float(parameters.get("state_mfi_median", 90116.0))
    notes.append(f"state_mfi_median: {state_mfi_median} (from parameters.yaml; ACS B19113 DE statewide)")

    # cycle_5_status: read from DSB output if present, else 'pending'
    cycle_5_status = "pending"
    if dsb_path.exists():
        try:
            dsb_payload = json.loads(dsb_path.read_text(encoding="utf-8"))
            cycle_5_status = dsb_payload.get("cycle_5_status", "pending")
        except Exception:  # noqa: BLE001
            pass
    manifest.cycle_5_status = cycle_5_status
    notes.append(f"cycle_5_status: {cycle_5_status}")

    return LoadedRaw(
        grantees=grantees,
        food_resources_raw=food_resources_raw,
        tracts=tracts,
        state_mfi_median=state_mfi_median,
        manifest=manifest,
        lila_geojson=lila_geojson,
        sd2_geojson=sd2_geojson,
        dart_routes_geojson=dart_routes_geojson,
        acs_demographics_csv=acs_demographics_csv,
        mhc_csv=mhc_csv,
        mmg_counties_gdf=mmg_counties_gdf,
        target_tracts_gdf=target_tracts_gdf,
        weights_bg_gdf=weights_bg_gdf,
        geo_stack_available=geo_stack,
        notes=notes,
    )


# ---------------------------------------------------------------------------
# DSB grantees → GranteeInput
# ---------------------------------------------------------------------------


_CATEGORY_DEFAULT = "supermarket"  # methodology v0.2.0 category enum default


def _load_dsb_grantees(path: Path) -> list[GranteeInput]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    raw = payload.get("grantees") or []
    out: list[GranteeInput] = []
    for r in raw:
        cycle = r.get("cycle")
        if cycle is None:
            continue
        amount = r.get("amount_usd")
        if amount is None:
            amount = 0.0
        out.append(
            GranteeInput(
                cycle=int(cycle),
                grantee=r.get("grantee") or "",
                amount_usd=float(amount),
                category=r.get("category") or _CATEGORY_DEFAULT,
                storefront_address=r.get("storefront_address"),
                zip_code=None,
                awarded_date=r.get("awarded_date"),
            )
        )
    return out


# ---------------------------------------------------------------------------
# ACS tract JSON → CSV
# ---------------------------------------------------------------------------


def _acs_tract_json_to_csv(raw: bytes) -> bytes:
    """Convert Census ACS 5-year tract-level JSON to a flat CSV.

    Census API returns [header_row, *data_rows]; we add a GEOID column
    constructed as state+county+tract.
    """
    payload = json.loads(raw.decode("utf-8"))
    if not payload:
        return b""
    header, *rows = payload
    out_header = ["GEOID", *header]
    out_rows = []
    state_idx = header.index("state")
    county_idx = header.index("county")
    tract_idx = header.index("tract")
    for row in rows:
        geoid = f"{row[state_idx]}{row[county_idx]}{row[tract_idx]}"
        out_rows.append([geoid, *row])
    lines = [",".join(str(c) for c in out_header)]
    for r in out_rows:
        lines.append(",".join(_csv_cell(c) for c in r))
    return ("\n".join(lines) + "\n").encode("utf-8")


def _csv_cell(value: Any) -> str:
    s = "" if value is None else str(value)
    if "," in s or '"' in s or "\n" in s:
        return '"' + s.replace('"', '""') + '"'
    return s


# ---------------------------------------------------------------------------
# ACS tract demographics → per-tract lookup (S+6)
# ---------------------------------------------------------------------------


# Census API sentinel values for missing/null data. Treat as None on read.
# Documented at https://www.census.gov/programs-surveys/acs/library/handbooks/jam-values.html
# Key sentinels: -666666666 (no sample), -999999999 (estimate unavailable),
# -222222222 (median falls outside published range), -888888888 (estimate
# not applicable), -333333333 (median falls below published range).
_ACS_SENTINELS = {
    -666666666,
    -999999999,
    -222222222,
    -888888888,
    -333333333,
    -555555555,
}


def _coerce_acs_value(raw: Any) -> Optional[float]:
    """Coerce a Census ACS cell to float, returning None for sentinels/junk."""
    if raw is None:
        return None
    try:
        val = float(raw)
    except (TypeError, ValueError):
        return None
    if int(val) in _ACS_SENTINELS:
        return None
    return val


def _load_acs_tract_demographics(raw: bytes) -> dict[str, dict[str, Optional[float]]]:
    """Parse the Census ACS tract-level JSON and return a per-tract lookup.

    Returns a dict keyed by tract GEOID (state+county+tract, 11 chars) with
    values containing computed `poverty_rate` (0..1 fraction) and `mfi`
    (median family income in USD).

    `poverty_rate` is computed as B17001_002E / B17001_001E (numerator /
    denominator). Returns None when the denominator is missing, zero, or
    sentinel-valued. `mfi` is read from B19113_001E and returns None for
    Census sentinel values. The downstream SB 254 stage treats None as
    "income data missing" and uses the other rule (whichever has data) or
    flags the tract as not low-income when both are None.

    Defensive against:
      - Empty payloads (returns {}).
      - Missing required columns (returns {} and lets caller warn).
      - Per-row coercion failures (skips the row).
    """
    if not raw:
        return {}
    try:
        payload = json.loads(raw.decode("utf-8"))
    except (ValueError, UnicodeDecodeError):
        return {}
    if not payload:
        return {}
    header, *rows = payload
    required = ("state", "county", "tract", "B17001_001E", "B17001_002E", "B19113_001E")
    try:
        idx = {col: header.index(col) for col in required}
    except ValueError:
        # Missing required column — bail; downstream stays None on poverty/mfi.
        return {}

    out: dict[str, dict[str, Optional[float]]] = {}
    for row in rows:
        try:
            geoid = f"{row[idx['state']]}{row[idx['county']]}{row[idx['tract']]}"
        except (IndexError, TypeError):
            continue
        if len(geoid) != 11:
            continue

        pov_denom = _coerce_acs_value(row[idx["B17001_001E"]])
        pov_numer = _coerce_acs_value(row[idx["B17001_002E"]])
        if pov_denom is not None and pov_denom > 0 and pov_numer is not None:
            poverty_rate: Optional[float] = pov_numer / pov_denom
        else:
            poverty_rate = None

        mfi_raw = _coerce_acs_value(row[idx["B19113_001E"]])
        # MFI of 0 is not meaningful; treat as missing.
        mfi: Optional[float] = mfi_raw if (mfi_raw is not None and mfi_raw > 0) else None

        out[geoid] = {"poverty_rate": poverty_rate, "mfi": mfi}
    return out


# ---------------------------------------------------------------------------
# Pass-through bytes loader
# ---------------------------------------------------------------------------


def _load_passthrough(
    path: Path, manifest: Manifest, name: str, notes: list[str]
) -> Optional[bytes]:
    if not path.exists():
        notes.append(f"{name}: {path.name} not present; pass-through empty")
        return None
    body = path.read_bytes()
    _record_source(manifest, name, path)
    notes.append(f"{name}: {len(body)} bytes loaded")
    return body


# ---------------------------------------------------------------------------
# Manifest synthesis from on-disk file metadata
# ---------------------------------------------------------------------------


def _record_source(manifest: Manifest, name: str, path: Path) -> None:
    body = path.read_bytes()
    manifest.add_source(
        name,
        SourceEntry(
            url=f"file://{path.resolve().as_posix()}",
            last_fetched=_mtime_iso(path),
            http_status=200,
            sha256=hashlib.sha256(body).hexdigest(),
            raw_path=str(path).replace("\\", "/"),
            size_bytes=len(body),
            warnings=[],
        ),
    )


def _mtime_iso(path: Path) -> str:
    ts = datetime.datetime.fromtimestamp(
        path.stat().st_mtime, tz=datetime.timezone.utc
    ).replace(microsecond=0)
    return ts.isoformat().replace("+00:00", "Z")


# ---------------------------------------------------------------------------
# SNAP retailers + farmers markets → food_resources_raw (S+5)
# ---------------------------------------------------------------------------


def _load_snap_retailers(path: Path) -> list[FoodResource]:
    """Read snap-retailers-de.geojson and produce FoodResource records."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    features = payload.get("features") or []
    out: list[FoodResource] = []
    for f in features:
        props = f.get("properties") or {}
        geom = f.get("geometry") or {}
        coords = geom.get("coordinates") or []
        if not coords or len(coords) < 2:
            continue
        lon, lat = float(coords[0]), float(coords[1])
        name = props.get("Store_Name") or ""
        if not name:
            continue
        store_type = props.get("Store_Type") or ""
        out.append(
            FoodResource(
                source="snap-retailers",
                name=str(name),
                lat=lat,
                lon=lon,
                category=_snap_store_type_to_category(store_type),
                address=_compose_snap_address(props),
                external_id=str(props.get("Record_ID") or "") or None,
            )
        )
    return out


_SNAP_STORE_TYPE_MAP = {
    "Supermarket": "supermarket",
    "Super Store": "supermarket",
    "Grocery Store": "grocery-store",
    "Specialty Store": "specialty-grocer",
    "Convenience Store": "corner-store",
    "Combination Grocery/Other": "grocery-store",
    "Farmers Market": "farmers-market",
    "Wholesaler": "wholesaler",
    "Bakery": "specialty-grocer",
    "Delivery Route": "other",
    "Military Commissary": "other",
}


def _snap_store_type_to_category(store_type: str) -> str:
    """Map USDA SNAP Store_Type values to methodology v0.2.0 categories.

    Unknown values fall through to 'other' which is excluded from the
    food-resource universe by the merge stage (food resources must be
    a recognized category to count toward SB 254 access).
    """
    return _SNAP_STORE_TYPE_MAP.get(store_type or "", "other")


def _compose_snap_address(props: dict) -> Optional[str]:
    parts = [
        props.get("Store_Street_Address"),
        props.get("City"),
        props.get("State"),
        props.get("Zip_Code"),
    ]
    cleaned = [str(p).strip() for p in parts if p]
    return ", ".join(cleaned) if cleaned else None


def _load_farmers_markets(path: Path) -> list[FoodResource]:
    """Read usda-farmers-markets-de.json and produce FoodResource records."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    markets = payload.get("markets") or []
    out: list[FoodResource] = []
    for m in markets:
        lat = m.get("location_lat")
        lon = m.get("location_lon")
        name = m.get("listing_name") or ""
        if not name or lat is None or lon is None:
            continue
        out.append(
            FoodResource(
                source="usda-farmers-markets",
                name=str(name),
                lat=float(lat),
                lon=float(lon),
                category="farmers-market",
                address=m.get("location_address"),
                external_id=str(m.get("listing_id") or "") or None,
            )
        )
    return out


# ---------------------------------------------------------------------------
# Geo-stack helpers (guarded by ImportError in the caller)
# ---------------------------------------------------------------------------


def _load_urban_areas_from_national_tiger(
    zip_path: Path, manifest: Manifest, name: str, notes: list[str]
) -> Optional[Any]:
    """Load the national TIGER UAC10 zip → GeoDataFrame of all 2010 urban areas.

    Methodology v0.3.0: we don't filter by state here. The downstream
    spatial join in `_compute_tract_urbanicity_lookup` clips per-tract by
    geometric intersection, so any UA polygon that touches a DE tract
    counts (this matters for cross-state UAs around Wilmington).
    """
    if not zip_path.exists():
        notes.append(
            f"{name}: {zip_path.name} not present; "
            "tract urbanicity will fall back to the county-level proxy"
        )
        return None
    try:
        import geopandas as gpd
    except ImportError:
        notes.append(f"{name}: geopandas unavailable; urban_areas_gdf=None")
        return None
    try:
        gdf = gpd.read_file(f"zip://{zip_path.resolve().as_posix()}")
    except Exception as exc:  # noqa: BLE001
        notes.append(f"{name}: read_file failed ({exc})")
        return None
    _record_source(manifest, name, zip_path)
    notes.append(f"{name}: read {len(gdf)} urban area(s) from {zip_path.name}")
    return gdf


def _compute_tract_urbanicity_lookup(
    target_tracts_gdf: Any,
    urban_areas_gdf: Any,
    notes: list[str],
) -> dict[str, str]:
    """For each DE tract, classify `urban` if it intersects any UAC10 polygon.

    Returns {tract_geoid: 'urban' | 'nonurban'} for every tract in the
    target GDF. The classification follows USDA LILA's rule: a tract is
    urban if any part of it falls within a 2010 Urban Area or Urban
    Cluster (UATYP10 ∈ {'U', 'C'}); else nonurban.

    A geopandas `sjoin(predicate='intersects')` does the heavy lifting.
    Results align with the LILA Food Access Research Atlas urban/rural
    test by construction (same UAC10 vintage).
    """
    geoid_col = None
    for candidate in ("GEOID", "GEOID20", "GEOID10"):
        if candidate in target_tracts_gdf.columns:
            geoid_col = candidate
            break
    if geoid_col is None:
        notes.append("tiger-uac: tract GDF has no GEOID column; cannot compute urbanicity lookup")
        return {}

    try:
        import geopandas as gpd
    except ImportError:  # pragma: no cover — already guarded upstream
        return {}

    # Match CRS before sjoin. TIGER ships in EPSG:4269 (NAD83); both layers
    # share that vintage so reproject defensively to the tract CRS.
    urban = urban_areas_gdf
    if (
        getattr(target_tracts_gdf, "crs", None) is not None
        and getattr(urban, "crs", None) is not None
        and target_tracts_gdf.crs != urban.crs
    ):
        urban = urban.to_crs(target_tracts_gdf.crs)

    # LILA counts both Urbanized Areas (UATYP10='U') and Urban Clusters
    # (UATYP10='C') as "urban". If the column is absent (some TIGER vintages
    # rename it), fall back to "every UA polygon counts".
    if "UATYP10" in urban.columns:
        urban = urban[urban["UATYP10"].isin(("U", "C"))]

    try:
        joined = gpd.sjoin(
            target_tracts_gdf[[geoid_col, "geometry"]],
            urban[["geometry"]],
            how="left",
            predicate="intersects",
        )
    except Exception as exc:  # noqa: BLE001
        notes.append(f"tiger-uac: sjoin failed ({exc}); falling back to county proxy")
        return {}

    # A tract is `urban` if it has at least one matching UA row.
    # index_right is NaN when no UA polygon intersected.
    urban_geoids = {
        str(row[geoid_col])
        for _, row in joined.iterrows()
        if _sjoin_has_match(row)
    }
    out: dict[str, str] = {}
    for _, row in target_tracts_gdf.iterrows():
        gid = str(row[geoid_col])
        out[gid] = "urban" if gid in urban_geoids else "nonurban"

    urban_count = sum(1 for v in out.values() if v == "urban")
    notes.append(
        f"tiger-uac: urbanicity lookup built — {urban_count}/{len(out)} "
        f"tract(s) classified urban (UAC10 intersect)"
    )
    return out


def _sjoin_has_match(row: Any) -> bool:
    """Return True when an sjoin row's index_right column is a real match."""
    try:
        val = row["index_right"]
    except (KeyError, TypeError):
        return False
    # NaN check that works for both numpy floats and None.
    return val == val and val is not None


def _load_de_counties_from_national_tiger(
    zip_path: Path, manifest: Manifest, name: str, notes: list[str]
) -> Optional[Any]:
    """Load the national TIGER counties zip and clip to Delaware (STATEFP='10')."""
    if not zip_path.exists():
        notes.append(f"{name}: {zip_path.name} not present; mmg_counties_gdf=None (apportionment skipped)")
        return None
    try:
        import geopandas as gpd
    except ImportError:
        notes.append(f"{name}: geopandas unavailable; mmg_counties_gdf=None")
        return None
    try:
        gdf = gpd.read_file(f"zip://{zip_path.resolve().as_posix()}")
    except Exception as exc:  # noqa: BLE001
        notes.append(f"{name}: read_file failed ({exc})")
        return None
    if "STATEFP" not in gdf.columns:
        notes.append(f"{name}: STATEFP column missing in TIGER county shapefile")
        return None
    de = gdf[gdf["STATEFP"] == "10"].copy()
    _record_source(manifest, name, zip_path)
    notes.append(f"{name}: clipped national counties to DE ({len(de)} feature(s))")
    return de


def _load_tiger_zip(
    zip_path: Path, manifest: Manifest, name: str, notes: list[str]
) -> Optional[Any]:
    if not zip_path.exists():
        notes.append(f"{name}: {zip_path.name} not present; GDF None")
        return None
    try:
        import geopandas as gpd
    except ImportError:
        notes.append(f"{name}: geopandas unavailable; skipping shapefile read")
        return None

    try:
        gdf = gpd.read_file(f"zip://{zip_path.resolve().as_posix()}")
    except Exception as exc:  # noqa: BLE001
        notes.append(f"{name}: read_file failed ({exc})")
        return None
    _record_source(manifest, name, zip_path)
    notes.append(f"{name}: read {len(gdf)} features from {zip_path.name}")
    return gdf


def _load_acs_bg_pop(path: Path) -> dict[str, int]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not payload:
        return {}
    header, *rows = payload
    pop_idx = header.index("B01003_001E")
    state_idx = header.index("state")
    county_idx = header.index("county")
    tract_idx = header.index("tract")
    bg_idx = header.index("block group")
    out: dict[str, int] = {}
    for row in rows:
        geoid = f"{row[state_idx]}{row[county_idx]}{row[tract_idx]}{row[bg_idx]}"
        try:
            out[geoid] = int(row[pop_idx])
        except (TypeError, ValueError):
            out[geoid] = 0
    return out


def _attach_bg_pop(weights_bg_gdf: Any, bg_pop_lookup: dict[str, int]) -> Any:
    """Add a POP column to the BG GeoDataFrame keyed by GEOID."""
    geoid_col = None
    for candidate in ("GEOID", "GEOID20", "GEOID10"):
        if candidate in weights_bg_gdf.columns:
            geoid_col = candidate
            break
    if geoid_col is None:
        return weights_bg_gdf
    weights_bg_gdf = weights_bg_gdf.copy()
    weights_bg_gdf["POP"] = weights_bg_gdf[geoid_col].map(bg_pop_lookup).fillna(0).astype(int)
    return weights_bg_gdf


def _build_tracts_from_gdfs(
    target_tracts_gdf: Any,
    weights_bg_gdf: Optional[Any],
    notes: list[str],
    demographics_lookup: Optional[dict[str, dict[str, Optional[float]]]] = None,
    urbanicity_lookup: Optional[dict[str, str]] = None,
) -> list[TractInput]:
    """Construct TractInput records (with BG centroids + pop + demographics) from GDFs.

    This is the SB 254-effective input shape. Without BG pop we still
    emit tracts with empty BG tuples — the SB 254 stage will then
    report population_total=0 for those tracts and skip them.

    When `demographics_lookup` is supplied (S+6), each tract's
    `poverty_rate` and `mfi` are joined from the ACS tract demographics.
    Tracts not present in the lookup keep poverty_rate=None and mfi=None,
    which the SB 254 stage flags as "income data missing".

    When `urbanicity_lookup` is supplied (methodology v0.3.0), each
    tract's `urbanicity` is read from the UAC10 cross-walk. When absent,
    the function falls back to the county-level proxy (FIPS 10003 →
    urban; others → nonurban) documented as a v0.2.x caveat.
    """
    demographics_lookup = demographics_lookup or {}
    urbanicity_lookup = urbanicity_lookup or {}
    geoid_col = None
    for candidate in ("GEOID", "GEOID20", "GEOID10"):
        if candidate in target_tracts_gdf.columns:
            geoid_col = candidate
            break
    if geoid_col is None:
        notes.append("tracts: no GEOID column on tract GDF; returning empty list")
        return []

    # Build BG lookup keyed by tract GEOID -> list[BlockGroup]
    bg_by_tract: dict[str, list[BlockGroup]] = {}
    if weights_bg_gdf is not None:
        bg_geoid_col = None
        for candidate in ("GEOID", "GEOID20", "GEOID10"):
            if candidate in weights_bg_gdf.columns:
                bg_geoid_col = candidate
                break
        if bg_geoid_col is not None:
            for _, row in weights_bg_gdf.iterrows():
                bg_id = str(row[bg_geoid_col])
                # BG GEOID is 12 chars (state2 + county3 + tract6 + bg1); tract GEOID is 11 chars
                tract_geoid = bg_id[:-1] if len(bg_id) == 12 else bg_id[:11]
                centroid = row.geometry.centroid if row.geometry is not None else None
                if centroid is None:
                    continue
                pop = int(row.get("POP", 0))
                bg_by_tract.setdefault(tract_geoid, []).append(
                    BlockGroup(
                        bg_geoid=bg_id,
                        centroid_lat=float(centroid.y),
                        centroid_lon=float(centroid.x),
                        population=pop,
                    )
                )

    # Urbanicity classification:
    #   - Methodology v0.3.0 (preferred): per-tract UAC10 cross-walk via
    #     `urbanicity_lookup`. Aligns with USDA LILA's own urban/rural test.
    #   - Methodology v0.2.x fallback (when UAC zip absent): county-level
    #     proxy — New Castle (FIPS 10003) → urban; Kent + Sussex → nonurban.
    #     Documented as a caveat in methodology v0.2.1.
    def _urbanicity(tract_geoid: str) -> str:
        if tract_geoid in urbanicity_lookup:
            return urbanicity_lookup[tract_geoid]
        return "urban" if tract_geoid.startswith("10003") else "nonurban"

    tracts: list[TractInput] = []
    for _, row in target_tracts_gdf.iterrows():
        tract_geoid = str(row[geoid_col])
        demos = demographics_lookup.get(tract_geoid, {})
        tracts.append(
            TractInput(
                tract_geoid=tract_geoid,
                urbanicity=_urbanicity(tract_geoid),
                block_groups=tuple(bg_by_tract.get(tract_geoid, [])),
                poverty_rate=demos.get("poverty_rate"),
                mfi=demos.get("mfi"),
            )
        )
    return tracts


# ---------------------------------------------------------------------------
# USDA LILA xlsx → DE-clipped GeoJSON (S+6 transform)
# ---------------------------------------------------------------------------


# Columns we surface from the USDA Food Access Research Atlas workbook.
# Names match the published 2019 release. The transform is defensive:
# columns missing from the workbook are skipped with a note rather than
# raising — USDA has renamed columns between vintages.
#
# Reference: USDA ERS Food Access Research Atlas Documentation,
# https://www.ers.usda.gov/data-products/food-access-research-atlas/documentation/
_LILA_REQUIRED_COL = "CensusTract"
_LILA_FLAG_COLS = (
    "LILATracts_1And10",
    "LILATracts_halfAnd10",
    "LILATracts_1And20",
    "LILATracts_Vehicle",
    "LowIncomeTracts",
    "LATracts1",
    "LATracts_half",
    "LATracts10",
    "LATracts20",
    "LATracts_Vehicle",
)
_LILA_NUMERIC_COLS = (
    "PovertyRate",
    "MedianFamilyIncome",
    "Urban",
    "POP2010",
    "OHU2010",
)


def _build_lila_geojson_from_xlsx(
    xlsx_path: Path,
    target_tracts_gdf: Any,
    notes: list[str],
) -> Optional[bytes]:
    """Parse USDA LILA xlsx, filter to DE, join TIGER geometry → GeoJSON bytes.

    Returns None when:
      - openpyxl is unavailable
      - the workbook has no recognizable data sheet
      - the header row lacks `CensusTract`
      - zero DE tracts match the TIGER geometry

    On success, writes a diagnostic note recording the join coverage and
    returns the serialized GeoJSON bytes (UTF-8). The output is a
    FeatureCollection whose features carry the LILA flags + selected
    numeric columns as properties.
    """
    try:
        import openpyxl  # type: ignore  noqa: F401
    except ImportError:
        notes.append("usda-lila-xlsx: openpyxl not installed; skipping LILA transform")
        return None

    try:
        records = _read_lila_xlsx(xlsx_path)
    except Exception as exc:  # noqa: BLE001
        notes.append(f"usda-lila-xlsx: workbook read failed ({type(exc).__name__}: {exc})")
        return None

    if not records:
        notes.append(f"usda-lila-xlsx: {xlsx_path.name} parsed but no DE tracts found")
        return None

    geoid_col = None
    for candidate in ("GEOID", "GEOID20", "GEOID10"):
        if candidate in target_tracts_gdf.columns:
            geoid_col = candidate
            break
    if geoid_col is None:
        notes.append("usda-lila-xlsx: target_tracts_gdf has no GEOID column; skipping LILA join")
        return None

    features: list[dict] = []
    join_hits = 0
    for _, row in target_tracts_gdf.iterrows():
        tract_geoid = str(row[geoid_col])
        attrs = records.get(tract_geoid)
        if attrs is None:
            continue
        geom = row.geometry
        if geom is None:
            continue
        try:
            geometry = json.loads(_shapely_to_geojson(geom))
        except Exception:  # noqa: BLE001
            continue
        join_hits += 1
        features.append(
            {
                "type": "Feature",
                "geometry": geometry,
                "properties": {"tract_geoid": tract_geoid, **attrs},
            }
        )

    if join_hits == 0:
        notes.append(
            f"usda-lila-xlsx: {len(records)} DE record(s) in workbook but "
            "zero matched TIGER GEOIDs; check FIPS column formatting"
        )
        return None

    notes.append(
        f"usda-lila-xlsx: built FeatureCollection — "
        f"{join_hits} tracts joined / {len(records)} DE records in workbook"
    )
    payload = {"type": "FeatureCollection", "features": features}
    return (json.dumps(payload, indent=2) + "\n").encode("utf-8")


def _shapely_to_geojson(geom: Any) -> str:
    """Serialize a shapely geometry to a GeoJSON-geometry JSON string.

    geopandas attaches `__geo_interface__` to shapely geoms; we json-dump
    that. Kept as a tiny helper so we can swap to shapely.to_geojson if
    we ever need GeoJSON-spec validation.
    """
    return json.dumps(geom.__geo_interface__)


def _read_lila_xlsx(xlsx_path: Path) -> dict[str, dict[str, Any]]:
    """Open the LILA workbook, find the data sheet, return {geoid: attrs} for DE.

    The workbook ships with a "Read Me" tab plus the data tab. We pick the
    first sheet whose header row contains `CensusTract`. The header row is
    not guaranteed to be row 1 — some vintages prepend a title row — so we
    scan the first 8 rows for the marker.

    Filters to Delaware (FIPS prefix "10") and only retains columns named
    in `_LILA_FLAG_COLS` / `_LILA_NUMERIC_COLS` (missing columns are
    silently skipped). Returns an empty dict when nothing matches.
    """
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        for sheet in wb.worksheets:
            data_row_offset, header = _find_lila_header(sheet)
            if header is None:
                continue
            tract_idx = header.index(_LILA_REQUIRED_COL)
            flag_lookup = {col: header.index(col) for col in _LILA_FLAG_COLS if col in header}
            numeric_lookup = {col: header.index(col) for col in _LILA_NUMERIC_COLS if col in header}
            out: dict[str, dict[str, Any]] = {}
            for row in sheet.iter_rows(min_row=data_row_offset + 1, values_only=True):
                if row is None or len(row) <= tract_idx:
                    continue
                raw_tract = row[tract_idx]
                geoid = _normalize_lila_tract(raw_tract)
                if geoid is None or not geoid.startswith("10"):
                    continue
                attrs: dict[str, Any] = {}
                for col, idx in flag_lookup.items():
                    if idx < len(row):
                        attrs[col] = _coerce_lila_flag(row[idx])
                for col, idx in numeric_lookup.items():
                    if idx < len(row):
                        attrs[col] = _coerce_lila_numeric(row[idx])
                out[geoid] = attrs
            if out:
                return out
        return {}
    finally:
        wb.close()


def _find_lila_header(sheet: Any, scan_rows: int = 8) -> tuple[int, Optional[list[str]]]:
    """Return (1-indexed header row, header values) for the first row containing CensusTract.

    Returns (0, None) when no such row is found in the first `scan_rows`.
    """
    for i, row in enumerate(sheet.iter_rows(max_row=scan_rows, values_only=True), start=1):
        if row is None:
            continue
        values = [str(c).strip() if c is not None else "" for c in row]
        if _LILA_REQUIRED_COL in values:
            return i, values
    return 0, None


def _normalize_lila_tract(value: Any) -> Optional[str]:
    """USDA's CensusTract column may be int or string; normalize to 11-char FIPS."""
    if value is None:
        return None
    if isinstance(value, float):
        # Tract FIPS shouldn't have decimals, but openpyxl can promote ints
        # when a column has any blank cells. Trust the int conversion.
        try:
            value = int(value)
        except (TypeError, ValueError):
            return None
    s = str(value).strip()
    if not s:
        return None
    # Zero-pad to 11; longer values get truncated to 11 (defensive).
    s = s.zfill(11)
    if len(s) > 11:
        return None
    return s


def _coerce_lila_flag(value: Any) -> int:
    """Coerce LILA flag cells to 0/1. Empty / non-numeric → 0."""
    if value is None or value == "":
        return 0
    try:
        return 1 if int(float(value)) >= 1 else 0
    except (TypeError, ValueError):
        return 0


def _coerce_lila_numeric(value: Any) -> Optional[float]:
    """Coerce numeric LILA columns to float. Empty / non-numeric → None."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# DART GTFS shapes.txt → dart-routes.geojson (S+6 transform)
# ---------------------------------------------------------------------------


# GTFS shapes.txt is a flat CSV: one row per (shape_id, sequence) point.
# Reconstructing LineStrings is plain CSV grouping — no geopandas needed.
# We also join trips.txt + routes.txt so each LineString carries the route
# short_name / long_name properties the dashboard map uses for labels.
#
# GTFS reference: https://gtfs.org/schedule/reference/


def _build_dart_routes_geojson_from_gtfs(
    zip_path: Path, notes: list[str]
) -> Optional[bytes]:
    """Read DART GTFS zip, reconstruct LineStrings from shapes.txt → GeoJSON bytes.

    Returns None when:
      - zip is unreadable or missing `shapes.txt`
      - shapes.txt has no usable rows

    On success, writes a diagnostic note recording the line count and
    returns the serialized GeoJSON bytes (UTF-8). Each Feature is a
    LineString with properties: `shape_id`, `route_short_names`,
    `route_long_names`. Trips/routes joins are best-effort — if those
    files are missing or malformed, the LineStrings still emit with
    empty route lists.
    """
    import zipfile

    try:
        with zipfile.ZipFile(zip_path) as zf:
            names = set(zf.namelist())
            if "shapes.txt" not in names:
                notes.append(f"dart-gtfs-zip: {zip_path.name} missing shapes.txt; skipping LineString build")
                return None
            shapes_csv = zf.read("shapes.txt").decode("utf-8-sig")
            trips_csv = zf.read("trips.txt").decode("utf-8-sig") if "trips.txt" in names else ""
            routes_csv = zf.read("routes.txt").decode("utf-8-sig") if "routes.txt" in names else ""
    except (zipfile.BadZipFile, KeyError, OSError) as exc:
        notes.append(f"dart-gtfs-zip: read failed ({type(exc).__name__}: {exc})")
        return None

    shape_lines = _build_shape_lines(shapes_csv)
    if not shape_lines:
        notes.append(f"dart-gtfs-zip: {zip_path.name} shapes.txt had no usable rows")
        return None

    routes_by_shape = _gtfs_routes_by_shape(trips_csv, routes_csv)

    features: list[dict] = []
    for shape_id, coords in shape_lines.items():
        # GTFS allows a shape with a single point, but a LineString needs ≥2.
        if len(coords) < 2:
            continue
        route_meta = routes_by_shape.get(shape_id, {"short_names": [], "long_names": []})
        features.append(
            {
                "type": "Feature",
                "geometry": {"type": "LineString", "coordinates": coords},
                "properties": {
                    "shape_id": shape_id,
                    "route_short_names": route_meta["short_names"],
                    "route_long_names": route_meta["long_names"],
                },
            }
        )

    if not features:
        notes.append(f"dart-gtfs-zip: {zip_path.name} produced 0 LineStrings (all shapes had <2 points)")
        return None

    notes.append(
        f"dart-gtfs-zip: built FeatureCollection — {len(features)} LineString(s) "
        f"from {len(shape_lines)} shape(s) in shapes.txt"
    )
    payload = {"type": "FeatureCollection", "features": features}
    return (json.dumps(payload, indent=2) + "\n").encode("utf-8")


def _build_shape_lines(shapes_csv: str) -> dict[str, list[list[float]]]:
    """Parse GTFS shapes.txt → {shape_id: [[lon, lat], ...]} sorted by sequence.

    Tolerates rows with malformed coords or sequence numbers; skips them.
    """
    import csv

    grouped: dict[str, list[tuple[int, float, float]]] = {}
    reader = csv.DictReader(io.StringIO(shapes_csv))
    for row in reader:
        try:
            shape_id = str(row["shape_id"])
            seq = int(row["shape_pt_sequence"])
            lat = float(row["shape_pt_lat"])
            lon = float(row["shape_pt_lon"])
        except (KeyError, ValueError, TypeError):
            continue
        grouped.setdefault(shape_id, []).append((seq, lon, lat))

    out: dict[str, list[list[float]]] = {}
    for shape_id, rows in grouped.items():
        rows.sort(key=lambda r: r[0])
        out[shape_id] = [[lon, lat] for _, lon, lat in rows]
    return out


def _gtfs_routes_by_shape(
    trips_csv: str, routes_csv: str
) -> dict[str, dict[str, list[str]]]:
    """Build {shape_id: {short_names, long_names}} from trips.txt + routes.txt.

    Multiple routes can share a shape; the lists are deduped while
    preserving first-seen order. Returns {} when either file is missing,
    malformed, or empty — the caller falls back to empty route lists.
    """
    import csv

    if not trips_csv or not routes_csv:
        return {}

    routes_by_id: dict[str, dict[str, str]] = {}
    try:
        for row in csv.DictReader(io.StringIO(routes_csv)):
            rid = str(row.get("route_id") or "")
            if not rid:
                continue
            routes_by_id[rid] = {
                "short_name": str(row.get("route_short_name") or ""),
                "long_name": str(row.get("route_long_name") or ""),
            }
    except (csv.Error, ValueError):
        return {}

    by_shape: dict[str, dict[str, list[str]]] = {}
    try:
        for row in csv.DictReader(io.StringIO(trips_csv)):
            shape_id = str(row.get("shape_id") or "")
            route_id = str(row.get("route_id") or "")
            if not shape_id or not route_id:
                continue
            route = routes_by_id.get(route_id)
            if route is None:
                continue
            entry = by_shape.setdefault(shape_id, {"short_names": [], "long_names": []})
            if route["short_name"] and route["short_name"] not in entry["short_names"]:
                entry["short_names"].append(route["short_name"])
            if route["long_name"] and route["long_name"] not in entry["long_names"]:
                entry["long_names"].append(route["long_name"])
    except (csv.Error, ValueError):
        return {}

    return by_shape


# ---------------------------------------------------------------------------
# parameters.yaml loader (lightweight; keeps run_etl import-clean)
# ---------------------------------------------------------------------------


def load_parameters(path: Path) -> dict:
    return yaml.safe_load(path.read_text(encoding="utf-8"))
